import pandas as pd
import sqlite3
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment
from sql import *

def rows_to_dataframe(rows):
    columns = ['num_train_samples', 'loss_bin_l', 'loss_bin_u', 'SUM(perfect_model_count)', 'AVG(test_acc)', 'AVG(perfect_models_percentage)', 'AVG(train_time)']
    df = pd.DataFrame(rows, columns=columns)
    return df

def get_results_df_by_model_name(model_name):
    db_path = os.path.join("output/table2/mnist_guess_rnn", "databases", f"{model_name}_stats.db")
    rows = get_model_stats_summary(db_path, verbose=False, return_print=False)
    df = rows_to_dataframe(rows)
    return df

def prepare_excel_sheet(excel_sheet):
    cols_titles = ["Arch", "Sample count", "Best Test Acc", "Train Loss (0.3,0.35)", "(0.35,0.4)", "(0.4,0.45)", "(0.45,0.5)", "(0.5,0.55)", "(0.55,0.6)", "(0.6,0.65)"]
    Architectures = ["Vanilla RNN", "Linear RNN", "Linear RNN complex diag_real_im", "Linear RNN complex stable_ring_init", "Linear RNN complex stable_ring_init_gamma"]
    sample_count = [16,8,4,2]
    cols = ["A","B","C","D","E","F","G","H","I","J"]
    for col_index,col in enumerate(cols):
        excel_sheet[f"{col}1"] = cols_titles[col_index]
    for row in range(2,22):
        index = row-2
        excel_sheet[f"B{str(row)}"] = sample_count[index%4]
        if index%4 == 0:
            excel_sheet[f"A{str(row)}"] = Architectures[int(index/4)]


def convert_df_to_dict(df, results_type):
    dict ={}
    for index in df.index:
        dict[(df['num_train_samples'][index],df['loss_bin_l'][index],df['loss_bin_u'][index])] = df[results_type][index]
        print(f"DEBUG: convert_df_to_dict: num_train_samples={df['num_train_samples'][index]}, loss_bin_l={df['loss_bin_l'][index]}, loss_bin_u={df['loss_bin_u'][index]}, {results_type}={df[results_type][index]}")
    print(f"convert_df_to_dict: dict: {dict}")
    return dict


def fill_arch_block_with_results(excel_sheet, arch_results_df, arch_index, results_type):
    cols = ["D", "E", "F", "G", "H", "I", "J"]
    samples_count = ["16", "8", "4", "2"]
    loss_bin_l = ["0.3","0.35","0.4","0.45","0.5","0.55","0.6"]
    loss_bin_u = ["0.35","0.4","0.45","0.5","0.55","0.6","0.65"]
    arch_results_dict = convert_df_to_dict(arch_results_df, results_type)
    resutls_best_value_per_sample_count_dict = get_resutls_best_value_per_sample_count_dict(arch_results_dict,samples_count,loss_bin_l,loss_bin_u)
    write_best_results_to_column_C = True
    for col_index,col in enumerate(cols): # Excel cols
        for index, sample_count in enumerate(samples_count): # Excel rows
            excel_row = arch_index*4 + 2 + index
            current_key = (int(sample_count),float(loss_bin_l[col_index]),float(loss_bin_u[col_index]))
            result_value = arch_results_dict.get(current_key)
            if result_value:
                excel_sheet[f"{col}{str(excel_row)}"] = result_value
                print(f"{results_type}={result_value}")
            else:
                excel_sheet[f"{col}{str(excel_row)}"] = "-"

            if write_best_results_to_column_C:
                excel_sheet[f"C{str(excel_row)}"] = resutls_best_value_per_sample_count_dict[sample_count]
                print(f"DEBUG: writing to sample_count={sample_count} in column C the value:{resutls_best_value_per_sample_count_dict[sample_count]}")
        write_best_results_to_column_C = False


def get_resutls_best_value_per_sample_count_dict(arch_results_dict,samples_count,loss_bin_l,loss_bin_u):
    resutls_best_value_per_sample_count_dict = {"2":0,"4":0,"8":0,"16":0}
    for sample_count in samples_count:
        for low, high in zip(loss_bin_l,loss_bin_u):
         current_combination = (int(sample_count), float(low), float(high))
         current_value = arch_results_dict.get(current_combination)
         if current_value and current_value>resutls_best_value_per_sample_count_dict[sample_count]:
            resutls_best_value_per_sample_count_dict[sample_count] = current_value

    return resutls_best_value_per_sample_count_dict

def resize_and_center_cells(excel_sheet):
    alignment = Alignment(horizontal='center', vertical='center')
    # Cells Alignment
    for row in excel_sheet.iter_rows():
        for cell in row:
            cell.alignment = alignment

    # Resizing cells
    for column in excel_sheet.columns:
        set_len = 0
        column_name = column[0].column_letter  # Get column name
        for cell in column:
            if len(str(cell.value)) > set_len:
                set_len = len(str(cell.value))
        set_col_width = set_len + 10
        excel_sheet.column_dimensions[column_name].width = set_col_width


if __name__ == "__main__":
    models_names = ["scale_0.1_N_32_H_in_1_edo_init_Inefficient_forward_pass_embeddings_type_linear_size_10_sigmoid_recurrent_rnn_same_seeds",
                    "scale_0.1_N_32_H_in_1_edo_init_Inefficient_forward_pass_embeddings_type_linear_size_10_linear_recurrent_rnn_same_seeds",
                    "scale_0.1_N_32_H_in_1_edo_init_Inefficient_forward_pass_embeddings_type_linear_size_10_linear_recurrent_complex_diag_real_im_parametrization_rnn_same_seeds",
                    "scale_0.1_N_32_H_in_1_edo_init_Inefficient_forward_pass_embeddings_type_linear_size_10_linear_recurrent_complex_diag_stable_ring_init_parametrization_r_min_0.0_r_max_1.0_max_phase_6.28_rnn_same_seeds",
                    "scale_0.1_N_32_H_in_1_edo_init_Inefficient_forward_pass_embeddings_type_linear_size_10_linear_recurrent_complex_diag_stable_ring_init_parametrization_r_min_0.0_r_max_1.0_max_phase_6.28_gamma_normalization_rnn_same_seeds"
                    ]

    model_name_parts = models_names[0].split('_')
    recurrent_index = model_name_parts.index('recurrent')
    excel_name_splitted = model_name_parts[0:recurrent_index-1] + model_name_parts[recurrent_index+1:]
    excel_name = "_".join(excel_name_splitted)
    workbook1 = openpyxl.Workbook()
    workbook2 = openpyxl.Workbook()
    test_acc_excel_sheet = workbook1.active
    perfect_models_percentage_excel_sheet = workbook2.active
    print(f"DEBUG: excel_name={excel_name}")
    for arch_index, arch_name in enumerate(models_names):
        current_arch_results_df = get_results_df_by_model_name(arch_name)
        prepare_excel_sheet(test_acc_excel_sheet)
        prepare_excel_sheet(perfect_models_percentage_excel_sheet)
        fill_arch_block_with_results(test_acc_excel_sheet, current_arch_results_df, arch_index, "AVG(test_acc)")
        fill_arch_block_with_results(perfect_models_percentage_excel_sheet, current_arch_results_df, arch_index, "AVG(perfect_models_percentage)")

    resize_and_center_cells(test_acc_excel_sheet)
    resize_and_center_cells(perfect_models_percentage_excel_sheet)
    tes_acc_excel_file_path = f'output/table2/mnist_guess_rnn/results_table_summary/{excel_name}_avg_test_accuracies.xlsx'
    perfect_models_percentage_excel_file_path = f'output/table2/mnist_guess_rnn/results_table_summary/{excel_name}_perfect_models_percentage.xlsx'
    workbook1.save(tes_acc_excel_file_path)
    workbook2.save(perfect_models_percentage_excel_file_path)


