import pandas as pd
import sqlite3
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment
from sql import *


def rows_to_dataframe(rows):
    columns = ['num_train_samples', 'loss_bin_l', 'loss_bin_u', 'SUM(perfect_model_count)', 'AVG(test_acc)', 'AVG(perfect_models_percentage)', 'AVG(train_time)']
    df = pd.DataFrame(rows, columns=columns)
    return df


# scale_0.1_N_32
# "scale_0.1_N_32_edo_init_Inefficient_forward_pass_linear_recurrent_complex_diag_stable_ring_init_parametrization_r_min_0.0_r_max_1.0_max_phase_0.31_rnn_same_seeds",
#models_names = ["scale_0.1_N_32_H_in_1_edo_init_Inefficient_forward_pass_embeddings_type_linear_size_10_sigmoid_recurrent_rnn_same_seeds",
#                "scale_0.1_N_32_H_in_1_edo_init_Inefficient_forward_pass_embeddings_type_linear_size_10_linear_recurrent_rnn_same_seeds"]
models_names= ["scale_0.1_N_32_H_in_1_edo_init_Inefficient_forward_pass_embeddings_type_linear_size_10_linear_recurrent_rnn_same_seeds"]
# scale_0.1_N_64
# models_names = ["scale_0.1_N_64_edo_init_Inefficient_forward_pass_sigmoid_recurrent_rnn_same_seeds",
#                 "scale_0.1_N_64_edo_init_Inefficient_forward_pass_linear_recurrent_rnn_same_seeds",
#                 "scale_0.1_N_64_edo_init_Inefficient_forward_pass_linear_recurrent_complex_diag_real_im_parametrization_rnn_same_seeds",
#                 "scale_0.1_N_64_edo_init_Inefficient_forward_pass_linear_recurrent_complex_diag_stable_ring_init_parametrization_r_min_0.0_r_max_1.0_max_phase_0.31_rnn_same_seeds",
#                 "scale_0.1_N_64_edo_init_Inefficient_forward_pass_linear_recurrent_complex_diag_stable_ring_init_parametrization_r_min_0.0_r_max_1.0_max_phase_0.31_gamma_normalization_rnn_same_seeds",
#                 "scale_0.1_N_64_edo_init_Inefficient_forward_pass_linear_recurrent_complex_diag_stable_ring_init_parametrization_r_min_0.0_r_max_1.0_max_phase_6.28_rnn_same_seeds",
#                 "scale_0.1_N_64_edo_init_Inefficient_forward_pass_linear_recurrent_complex_diag_stable_ring_init_parametrization_r_min_0.0_r_max_1.0_max_phase_6.28_gamma_normalization_rnn_same_seeds",
#                 "scale_0.1_N_64_edo_init_Inefficient_forward_pass_linear_recurrent_complex_diag_stable_ring_init_parametrization_r_min_0.8_r_max_1.0_max_phase_0.31_rnn_same_seeds",
#                 "scale_0.1_N_64_edo_init_Inefficient_forward_pass_linear_recurrent_complex_diag_stable_ring_init_parametrization_r_min_0.8_r_max_1.0_max_phase_0.31_gamma_normalization_rnn_same_seeds"
#                 ]
# scale_0.11_N_32_pix_to_vec_to_vec
# models_names = [ "scale_0.11_N_32_H_in_100_edo_init_Inefficient_forward_pass_embeddings_type_pix_to_vec_to_vec_size_10_sigmoid_recurrent_rnn_same_seeds",
#                  "scale_0.11_N_32_H_in_100_edo_init_Inefficient_forward_pass_embeddings_type_pix_to_vec_to_vec_size_10_linear_recurrent_rnn_same_seeds",
#                  "scale_0.11_N_32_H_in_100_edo_init_Inefficient_forward_pass_embeddings_type_pix_to_vec_to_vec_size_10_linear_recurrent_complex_diag_real_im_parametrization_rnn_same_seeds",
#                  "scale_0.11_N_32_H_in_100_edo_init_Inefficient_forward_pass_embeddings_type_pix_to_vec_to_vec_size_10_linear_recurrent_complex_diag_stable_ring_init_parametrization_r_min_0.0_r_max_1.0_max_phase_6.28_rnn_same_seeds",
#                  "scale_0.11_N_32_H_in_100_edo_init_Inefficient_forward_pass_embeddings_type_pix_to_vec_to_vec_size_10_linear_recurrent_complex_diag_stable_ring_init_parametrization_r_min_0.0_r_max_1.0_max_phase_6.28_gamma_normalization_rnn_same_seeds"
#                 ]


num_train_samples_list = [16,8,4,2]
arch_best_test_acc_df_dict = {}
arch_best_perfect_models_percentage_df_dict = {}
for model_name in models_names:
    db_path = os.path.join("output/table2/mnist_guess_rnn", "databases", f"{model_name}_stats.db")
    #print(f"DEBUG: current db_path={db_path}")
    print(f"DEBUG: current model_name={model_name}")
    rows = get_model_stats_summary(db_path, verbose=False, return_print=False)
    df = rows_to_dataframe(rows)
    #print(f"df:")
    #print(df)
    best_test_accuracy_per_num_train_sample_df = df.groupby('num_train_samples')['AVG(test_acc)'].max().reset_index()
    best_perfect_models_percentage_per_num_train_sample_df = df.groupby('num_train_samples')['AVG(perfect_models_percentage)'].max().reset_index()
    #print(f"best_test_accuracy_per_num_train_sample_df:")
    #print(best_test_accuracy_per_num_train_sample_df)
    arch_best_test_acc_df_dict[model_name] = best_test_accuracy_per_num_train_sample_df
    arch_best_perfect_models_percentage_df_dict[model_name] = best_perfect_models_percentage_per_num_train_sample_df

columns = ['num_train_samples','arch','best_AVG(test_acc)']
best_test_accuracies_and_perfect_models_percentage_summary_df = pd.DataFrame(columns=columns)
for num_train_samples in num_train_samples_list:
    for model_name in models_names:
        #print(f"DEBUG: num_train_samples={num_train_samples} model_name={model_name}")
        curr_best_test_acc_df = arch_best_test_acc_df_dict[model_name]
        curr_best_perfect_models_percentage_df = arch_best_perfect_models_percentage_df_dict[model_name]
        #print(f"curr_df")
        #print(curr_df)
        if not curr_best_test_acc_df.empty and (num_train_samples in curr_best_test_acc_df['num_train_samples'].values):
            best_test_acc_df = curr_best_test_acc_df.loc[(curr_best_test_acc_df['num_train_samples'] == num_train_samples)].reset_index()
            best_perfect_models_percentage_df = curr_best_perfect_models_percentage_df.loc[(curr_best_perfect_models_percentage_df['num_train_samples'] == num_train_samples)].reset_index()
            best_test_acc = best_test_acc_df.loc[0, "AVG(test_acc)"]
            best_perfect_models_percentage = best_perfect_models_percentage_df.loc[0, "AVG(perfect_models_percentage)"]
        else:
            best_test_acc = -1
            best_perfect_models_percentage = -1



        #print(f"DEBUG: after")
        data_point = {'num_train_samples': [num_train_samples], 'arch': [model_name], 'best_AVG(test_acc)': [best_test_acc], 'best_AVG(perfect_models_percentage)': [best_perfect_models_percentage]}
        data_point_df = pd.DataFrame(data_point)
        best_test_accuracies_and_perfect_models_percentage_summary_df = pd.concat([best_test_accuracies_and_perfect_models_percentage_summary_df,data_point_df],ignore_index=True)

experiments_prefix_parts = models_names[0].split('_')[0:4]
print(f"DEBUG:experiments_prefix_parts={experiments_prefix_parts}")
experiments_prefix = '_'.join(experiments_prefix_parts)
excel_file_path = f'output/table2/mnist_guess_rnn/best_test_acc_and_perfect_models_percentage_summary/{experiments_prefix}.xlsx'
print(f"DEBUG:excel_file_path={excel_file_path}")
print("Best test accuracies and perfect_models_percentage summary:")
print(best_test_accuracies_and_perfect_models_percentage_summary_df)
best_test_accuracies_and_perfect_models_percentage_summary_df.to_excel(excel_file_path, sheet_name='best_test_accuracies_and_perfect_models_percentage_summary', index=False)

# Handling Excel file
book = load_workbook(excel_file_path)
sheet = book.active
alignment = Alignment(horizontal='center', vertical='center')
# Cells Alignment
for row in sheet.iter_rows():
    for cell in row:
        cell.alignment = alignment

# Resizing cells
for column in sheet.columns:
    set_len = 0
    column_name = column[0].column_letter # Get column name
    for cell in column:
        if len(str(cell.value)) > set_len:
            set_len = len(str(cell.value))

    set_col_width = set_len + 5
    sheet.column_dimensions[column_name].width = set_col_width

# Rows merging in the first column
for row in range(2, sheet.max_row + 1, len(models_names)):
    cell_range = f'A{row}:A{row + len(models_names) - 1}'  # Define the range of cells to merge
    sheet.merge_cells(cell_range)  # Merge the cells
    merged_cell = sheet[f'A{row}']  # Get the merged cell
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')  # Apply alignment

book.save(excel_file_path)








